#!/usr/bin/env python3
"""
deploy_survey.py

End-to-end deployment: XLSForm -> GeoPackage + QGIS project -> Mergin Maps

Configuration precedence (highest to lowest):
  1. Command line arguments
  2. Config file (YAML)
  3. Environment variables (MERGIN_USERNAME, MERGIN_PASSWORD, MERGIN_URL)
  4. Interactive prompts (if running in a terminal)

Usage examples:

  Fully automated (config + env vars):
    MERGIN_USERNAME=jane MERGIN_PASSWORD=s3cret python deploy_survey.py --config deploy.yml

  CLI one-off:
    python deploy_survey.py --xlsform survey.xlsx --workspace my-org

  Minimal (will prompt for missing values):
    python deploy_survey.py --xlsform survey.xlsx
"""

import argparse
import getpass
import os
import shutil
import subprocess
import sys
import tempfile

try:
    import yaml
except ImportError:
    yaml = None

try:
    import mergin
    from mergin import MerginClient, ClientError
except ImportError:
    mergin = None

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ---------------------------------------------------------------------------
# Configuration resolution
# ---------------------------------------------------------------------------

def load_config_file(path):
    """Load a YAML config file and return a flat dict of resolved values."""
    if yaml is None:
        print("Error: PyYAML is required to use config files (pip install pyyaml)")
        sys.exit(1)

    with open(path, "r") as f:
        raw = yaml.safe_load(f) or {}

    mergin_block = raw.get("mergin", {}) or {}

    return {
        "url": mergin_block.get("url"),
        "username": mergin_block.get("username"),
        "password": mergin_block.get("password"),
        "workspace": raw.get("workspace"),
        "xlsform": raw.get("xlsform"),
        "project_name": raw.get("project_name"),
        "overwrite": raw.get("overwrite"),
    }


def parse_cli_args():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="Deploy an XLSForm survey to Mergin Maps",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Configuration precedence (highest to lowest):
  1. Command line arguments
  2. Config file (--config)
  3. Environment variables (MERGIN_USERNAME, MERGIN_PASSWORD, MERGIN_URL)
  4. Interactive prompts (when running in a terminal)

Examples:
  %(prog)s --config deploy.yml
  %(prog)s --xlsform survey.xlsx --workspace my-org
  %(prog)s --xlsform survey.xlsx --workspace my-org --username jane
        """,
    )
    parser.add_argument(
        "--config", "-c",
        help="Path to YAML config file",
    )
    parser.add_argument(
        "--xlsform", "-f",
        help="Path to XLSForm .xlsx file",
    )
    parser.add_argument(
        "--workspace", "-w",
        help="Mergin Maps workspace name",
    )
    parser.add_argument(
        "--project-name", "-p",
        help="Project name (default: inferred from XLSForm settings)",
    )
    parser.add_argument(
        "--url",
        help="Mergin Maps server URL (default: https://app.merginmaps.com)",
    )
    parser.add_argument(
        "--username", "-u",
        help="Mergin Maps username (default: MERGIN_USERNAME env var)",
    )
    parser.add_argument(
        "--password",
        help="Mergin Maps password (default: MERGIN_PASSWORD env var)",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        default=None,
        help="Overwrite if project already exists on Mergin Maps",
    )
    parser.add_argument(
        "--no-interactive",
        action="store_true",
        help="Disable interactive prompts (fail if required values are missing)",
    )
    parser.add_argument(
        "--converter",
        default="xlsform_to_gpkg.py",
        help="Path to the xlsform_to_gpkg.py converter script (default: xlsform_to_gpkg.py)",
    )

    return parser.parse_args()


def get_env_values():
    """Read configuration from environment variables."""
    return {
        "url": os.environ.get("MERGIN_URL"),
        "username": os.environ.get("MERGIN_USERNAME"),
        "password": os.environ.get("MERGIN_PASSWORD"),
    }


def prompt_for_missing(config, interactive):
    """Prompt for any required values still missing. Returns updated config."""
    if not interactive:
        return config

    if not sys.stdin.isatty():
        return config

    if not config.get("workspace"):
        config["workspace"] = input("Mergin Maps workspace: ").strip()

    if not config.get("xlsform"):
        config["xlsform"] = input("Path to XLSForm (.xlsx): ").strip()

    if not config.get("username"):
        config["username"] = input("Mergin Maps username: ").strip()

    if not config.get("password"):
        config["password"] = getpass.getpass("Mergin Maps password: ")

    return config


def resolve_config(args):
    """
    Merge all configuration sources in precedence order:
      CLI args > config file > env vars > interactive prompts
    """
    # Start with defaults
    config = {
        "url": "https://app.merginmaps.com",
        "username": None,
        "password": None,
        "workspace": None,
        "xlsform": None,
        "project_name": None,
        "overwrite": False,
        "converter": "xlsform_to_gpkg.py",
    }

    # Layer 1: environment variables
    env = get_env_values()
    for key, val in env.items():
        if val is not None:
            config[key] = val

    # Layer 2: config file
    if args.config:
        if not os.path.isfile(args.config):
            print(f"Error: config file not found: {args.config}")
            sys.exit(1)
        file_vals = load_config_file(args.config)
        for key, val in file_vals.items():
            if val is not None:
                config[key] = val

    # Layer 3: command line args (highest precedence)
    cli_map = {
        "url": args.url,
        "username": args.username,
        "password": args.password,
        "workspace": args.workspace,
        "xlsform": args.xlsform,
        "project_name": args.project_name,
        "overwrite": args.overwrite,
        "converter": args.converter,
    }
    for key, val in cli_map.items():
        if val is not None:
            config[key] = val

    # Layer 4: interactive prompts for anything still missing
    interactive = not args.no_interactive
    config = prompt_for_missing(config, interactive)

    return config


def validate_config(config):
    """Ensure all required configuration values are present."""
    missing = []
    if not config.get("xlsform"):
        missing.append("xlsform")
    if not config.get("workspace"):
        missing.append("workspace")
    if not config.get("username"):
        missing.append("username (set MERGIN_USERNAME or use --username)")
    if not config.get("password"):
        missing.append("password (set MERGIN_PASSWORD or use --password)")

    if missing:
        print("Error: missing required configuration:")
        for m in missing:
            print(f"  - {m}")
        sys.exit(1)

    if not os.path.isfile(config["xlsform"]):
        print(f"Error: XLSForm not found: {config['xlsform']}")
        sys.exit(1)

    if not os.path.isfile(config["converter"]):
        print(f"Error: converter script not found: {config['converter']}")
        sys.exit(1)


# ---------------------------------------------------------------------------
# XLSForm introspection
# ---------------------------------------------------------------------------

def infer_project_name(xlsform_path):
    """
    Try to read the project name from the XLSForm settings sheet.
    Falls back to the XLSForm filename (without extension).
    """
    if openpyxl is not None:
        try:
            wb = openpyxl.load_workbook(xlsform_path, read_only=True)
            if "settings" in wb.sheetnames:
                ws = wb["settings"]
                headers = [
                    cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))
                ]
                # Look for form_id or form_title in the settings sheet
                for target_col in ("form_id", "form_title"):
                    if target_col in headers:
                        col_idx = headers.index(target_col)
                        row = next(ws.iter_rows(min_row=2, max_row=2), None)
                        if row and row[col_idx].value:
                            name = str(row[col_idx].value).strip()
                            wb.close()
                            # Sanitize: lowercase, underscores for spaces
                            return name.replace(" ", "_").lower()
            wb.close()
        except Exception:
            pass  # Fall through to filename-based inference

    # Fallback: use the xlsx filename
    base = os.path.basename(xlsform_path)
    name, _ = os.path.splitext(base)
    return name.replace(" ", "_").lower()


# ---------------------------------------------------------------------------
# Deployment workflow
# ---------------------------------------------------------------------------

def generate_project(config, project_dir):
    """
    Run xlsform_to_gpkg.py to produce the .gpkg and .qgs files
    in the target project directory.
    """
    project_name = config["project_name"]
    gpkg_path = os.path.join(project_dir, f"{project_name}.gpkg")

    print(f"Generating project from: {config['xlsform']}")
    result = subprocess.run(
        [
            sys.executable,
            config["converter"],
            config["xlsform"],
            gpkg_path,
            "--overwrite",
        ],
        capture_output=True,
        text=True,
    )

    if result.returncode != 0:
        print("Error: xlsform_to_gpkg.py failed:")
        print(result.stderr)
        sys.exit(1)

    # Verify expected outputs exist
    qgs_path = os.path.join(project_dir, f"{project_name}.qgs")
    if not os.path.isfile(gpkg_path):
        print(f"Error: expected GeoPackage not found: {gpkg_path}")
        sys.exit(1)
    if not os.path.isfile(qgs_path):
        print(f"Error: expected QGIS project not found: {qgs_path}")
        sys.exit(1)

    print(f"  Created: {os.path.basename(gpkg_path)}")
    print(f"  Created: {os.path.basename(qgs_path)}")
    return gpkg_path, qgs_path


def deploy_to_mergin(config, project_dir):
    """
    Create the project on Mergin Maps and push the generated files.
    """
    if mergin is None:
        print("Error: mergin-client is required (pip install mergin-client)")
        sys.exit(1)

    full_project_name = f"{config['workspace']}/{config['project_name']}"
    print(f"Deploying to Mergin Maps: {full_project_name}")
    print(f"  Server: {config['url']}")

    # Authenticate
    try:
        client = MerginClient(
            url=config["url"],
            login=config["username"],
            password=config["password"],
        )
    except Exception as e:
        print(f"Error: authentication failed: {e}")
        sys.exit(1)

    print(f"  Authenticated as: {config['username']}")

    # Check if project already exists
    project_exists = False
    try:
        client.project_info(full_project_name)
        project_exists = True
    except ClientError:
        pass

    if project_exists:
        if not config.get("overwrite"):
            print(f"Error: project '{full_project_name}' already exists.")
            print("  Use --overwrite to replace it, or choose a different project name.")
            sys.exit(1)
        else:
            print(f"  Project exists — overwrite enabled, removing old project...")
            try:
                client.delete_project(full_project_name)
            except ClientError as e:
                print(f"Error: could not delete existing project: {e}")
                sys.exit(1)

    # Create the project on Mergin Maps
    try:
        client.create_project(full_project_name)
        print(f"  Created project: {full_project_name}")
    except ClientError as e:
        print(f"Error: could not create project: {e}")
        sys.exit(1)

    # Download the empty project to initialize .mergin sync metadata
    # We use a temporary directory to get the metadata, then move it
    # into our project directory
    try:
        # Download into the project_dir — if files already exist, the client
        # should initialize .mergin alongside them
        # First, move our generated files to a temp location
        temp_hold = tempfile.mkdtemp()
        generated_files = []
        for fname in os.listdir(project_dir):
            src = os.path.join(project_dir, fname)
            dst = os.path.join(temp_hold, fname)
            shutil.move(src, dst)
            generated_files.append(fname)

        # Download the empty project (creates .mergin metadata)
        client.download_project(full_project_name, project_dir)

        # Move generated files back
        for fname in generated_files:
            src = os.path.join(temp_hold, fname)
            dst = os.path.join(project_dir, fname)
            shutil.move(src, dst)

        shutil.rmtree(temp_hold)
        print("  Initialized sync metadata")

    except ClientError as e:
        print(f"Error: could not initialize project locally: {e}")
        sys.exit(1)

    # Push the project files
    try:
        client.push_project(project_dir)
        print(f"  Pushed project files to Mergin Maps")
    except ClientError as e:
        print(f"Error: push failed: {e}")
        sys.exit(1)

    print()
    print(f"Done! Project is live at: {config['url']}/projects/{full_project_name}")
    print("Team members can now sync this project to their devices.")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = parse_cli_args()
    config = resolve_config(args)

    # Infer project name if not provided
    if not config.get("project_name"):
        config["project_name"] = infer_project_name(config["xlsform"])
        print(f"Inferred project name: {config['project_name']}")

    validate_config(config)

    # Create a working directory for the project files
    project_dir = tempfile.mkdtemp(prefix="mergin_deploy_")
    print(f"Working directory: {project_dir}")

    try:
        # Step 1: Generate .gpkg and .qgs from XLSForm
        generate_project(config, project_dir)

        # Step 2: Push to Mergin Maps
        deploy_to_mergin(config, project_dir)

    except KeyboardInterrupt:
        print("\nAborted.")
        sys.exit(1)
    finally:
        # Clean up working directory
        if os.path.isdir(project_dir):
            shutil.rmtree(project_dir)


if __name__ == "__main__":
    main()